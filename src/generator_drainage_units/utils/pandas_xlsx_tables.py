from itertools import chain, cycle
from typing import BinaryIO, Iterable, Literal, Optional, Union

import numpy as np
import pandas as pd
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from openpyxl.worksheet.table import Table, TableStyleInfo
from pandas.api.types import (
    is_datetime64_any_dtype,
    is_float_dtype,
    is_integer_dtype,
    is_string_dtype,
)
from pandas.core.dtypes.common import is_list_like
from pandas.io.formats.format import FloatArrayFormatter

NamedTableStyle = Literal[
    "Table Style Dark 1",
    "Table Style Dark 2",
    "Table Style Dark 3",
    "Table Style Dark 4",
    "Table Style Dark 5",
    "Table Style Dark 6",
    "Table Style Dark 7",
    "Table Style Dark 8",
    "Table Style Dark 9",
    "Table Style Dark 10",
    "Table Style Dark 11",
    "Table Style Light 1",
    "Table Style Light 2",
    "Table Style Light 3",
    "Table Style Light 4",
    "Table Style Light 5",
    "Table Style Light 6",
    "Table Style Light 7",
    "Table Style Light 8",
    "Table Style Light 9",
    "Table Style Light 10",
    "Table Style Light 11",
    "Table Style Light 12",
    "Table Style Light 13",
    "Table Style Light 14",
    "Table Style Light 15",
    "Table Style Light 16",
    "Table Style Light 17",
    "Table Style Light 18",
    "Table Style Light 19",
    "Table Style Light 20",
    "Table Style Light 21",
    "Table Style Medium 1",
    "Table Style Medium 2",
    "Table Style Medium 3",
    "Table Style Medium 4",
    "Table Style Medium 5",
    "Table Style Medium 6",
    "Table Style Medium 7",
    "Table Style Medium 8",
    "Table Style Medium 9",
    "Table Style Medium 10",
    "Table Style Medium 11",
    "Table Style Medium 12",
    "Table Style Medium 13",
    "Table Style Medium 14",
    "Table Style Medium 15",
    "Table Style Medium 16",
    "Table Style Medium 17",
    "Table Style Medium 18",
    "Table Style Medium 19",
    "Table Style Medium 20",
    "Table Style Medium 21",
    "Table Style Medium 22",
    "Table Style Medium 23",
    "Table Style Medium 24",
    "Table Style Medium 25",
    "Table Style Medium 26",
    "Table Style Medium 27",
    "Table Style Medium 28",
]


def create_format_mapping(workbook):
    return {
        "text": workbook.add_format({"num_format": "@"}),
        "float1": workbook.add_format({"num_format": "0.0"}),
        "float2": workbook.add_format({"num_format": "0.00"}),
        "float3": workbook.add_format({"num_format": "0.000"}),
        "float6": workbook.add_format({"num_format": "0.000000"}),
        "int": workbook.add_format({"num_format": "0"}),
        "datetime-milliseconds": workbook.add_format(
            {"num_format": "yyyy-mm-dd hh:mm:ss.000"}
        ),
        "datetime-seconds": workbook.add_format({"num_format": "yyyy-mm-dd hh:mm:ss"}),
        "datetime-minutes": workbook.add_format({"num_format": "yyyy-mm-dd hh:mm"}),
        "date": workbook.add_format({"num_format": "yyyy-mm-dd"}),
        "scientific": workbook.add_format({"num_format": "0,00E+00"}),
    }


def format_for_col(col: pd.Series, format_mapping: dict):
    # https://pandas.pydata.org/pandas-docs/stable/user_guide/timeseries.html#offset-aliases
    if is_integer_dtype(col):
        return format_mapping["int"]

    elif is_float_dtype(col):
        # use pandas internal function to determine float format (number of
        # digits or scientific notation)
        suffix = (
            FloatArrayFormatter(col[~col.isna()]).get_result_as_array()[0].split(".")[1]
        )
        if "e" in suffix:
            return format_mapping["scientific"]
        digits = len(suffix)
        if digits <= 1:
            return format_mapping["float1"]
        elif digits <= 2:
            return format_mapping["float2"]
        elif digits <= 3:
            return format_mapping["float3"]
        else:
            return format_mapping["float6"]

    elif is_string_dtype(col):
        return format_mapping["text"]
    elif is_datetime64_any_dtype(col):
        # check smallest used time unit for precision
        if all((col - col.dt.floor("1D")) == pd.Timedelta(0)):
            return format_mapping["date"]
        elif all((col - col.dt.floor("min")) == pd.Timedelta(0)):
            return format_mapping["datetime-minutes"]
        elif all((col - col.dt.floor("S")) == pd.Timedelta(0)):
            return format_mapping["datetime-seconds"]
        else:
            return format_mapping["datetime-milliseconds"]


class TableNotFound(Exception):
    pass


def table_to_df(
    ws: ReadOnlyWorksheet,
    table: Table,
    index,
    values_as_nan={"#NUM!", "#VALUE!", "#N/A", "#NAME?", "#REF!", "#NULL!"},
    values_as_inf={"#DIV/0!"},
    values_as_empty_string={None},
) -> pd.DataFrame:
    columns = [col.name for col in table.tableColumns]
    data_rows = ws[table.ref][
        (table.headerRowCount or 0) : -table.totalsRowCount
        if table.totalsRowCount is not None
        else None
    ]
    data = ((cell.value for cell in row) for row in data_rows)
    replacements = chain(
        zip(values_as_empty_string, cycle([""])),
        zip(values_as_nan, cycle([np.nan])),
        zip(values_as_inf, cycle([np.inf])),
    )
    frame = pd.DataFrame(data, columns=columns, index=None)
    dtypes = frame.dtypes
    frame = frame.replace({k: r for k, r in replacements})
    for col, dtype in zip(frame.columns, dtypes):
        frame[col] = frame[col].astype(dtype)

    if index:
        if index == "auto":
            if table.tableStyleInfo.showFirstColumn:
                frame = frame.set_index(columns[0])
        elif index is False:
            pass
        elif is_list_like(index):
            frame = frame.set_index([columns[i] for i in index])
        else:
            frame = frame.set_index(columns[index])
    return frame


def xlsx_tables_to_dfs(
    file, index: [Literal["auto"], int, Iterable[int]] = "auto"
):
    """Get all tables from a given workbook. Returns a dictionary of tables.
    Requires a filename, which includes the file path and filename.

    Inspired by:
    https://github.com/pandas-dev/pandas/issues/24862#issuecomment-458885960
    https://stackoverflow.com/questions/43941365/openpyxl-read-tables-from-existing-data-book-example
    """

    # Load the workbook, from the filename, setting read_only to False
    wb = load_workbook(
        filename=file, read_only=False, keep_vba=False, data_only=True, keep_links=False
    )

    # Initialize the dictionary of tables
    return {
        name: table_to_df(ws, tbl, index)
        for ws in wb.worksheets
        for name, tbl in {**ws.tables}.items()
    }


def xlsx_table_to_df(
    file, table: str, index: Union[Literal["auto"], int, Iterable[int]] = "auto"
):
    """Get a table from a given workbook by the tablename."""

    # Load the workbook, from the filename, setting read_only to False
    wb = load_workbook(
        filename=file, read_only=False, keep_vba=False, data_only=True, keep_links=False
    )

    # Initialize the dictionary of tables

    for ws in wb.worksheets:
        if table in ws.tables:
            return table_to_df(ws, ws.tables[table], index)
    all_tables = {f"'{table}'" for ws in wb.worksheets for table in ws.tables.keys()}
    raise TableNotFound(
        f"Table '{table}' could not be found in the workbook. "
        f"Choose from {', '.join(all_tables)}."
    )


HeaderOrientation = Literal["diagonal", "horizontal", "vertical"]


def dfs_to_xlsx_tables(
    input: Iterable[tuple[pd.DataFrame, str]],
    file: Union[str, BinaryIO],
    index: bool = True,
    table_style: Optional[NamedTableStyle] = "Table Style Medium 9",
    nan_inf_to_errors=False,
    header_orientation: HeaderOrientation = "horizontal",
    remove_timezone: bool = False,
) -> None:
    """Convert multiple dataframes to an excel file.

    Args:
        input (Iterable[tuple[DataFrame, str]]): A list of tuples of (df, table_name)
        file (Union[str, BinaryIO]): File name or descriptor for the output
        index (bool, optional): Include the datafrme index in the results. Defaults
            to True
        table_style (Optional[NamedTableStyle], optional): Excel table style. Defaults
            to "Table Style Medium 9".
        nan_inf_to_errors (bool, optional): Explicitly write nan/inf values as errors.
            Defaults to False.
        header_orientation (HeaderOrientation, optional): Rotate the table headers, can
            be horizontal, vertical or diagonal. Defaults to "horizontal".
    """
    wb = xlsxwriter.Workbook(
        file,
        options=dict(
            nan_inf_to_errors=nan_inf_to_errors,
            remove_timezone=remove_timezone,
        ),
    )

    format_mapping = create_format_mapping(wb)
    if header_orientation == "diagonal":
        header_format = wb.add_format()
        header_format.set_rotation(45)
    elif header_orientation == "vertical":
        header_format = wb.add_format()
        header_format.set_rotation(90)

    for df, table_name in input:
        ws = wb.add_worksheet(name=table_name)
        if index:
            df = df.reset_index()
        if not nan_inf_to_errors:
            df = (
                df.replace(np.inf, np.finfo(np.float64).max)
                .replace(-np.inf, np.finfo(np.float64).min)
                .fillna("")
            )

        column_names = (str(c) for c in df.columns)
        options = {
            "data": df.values,
            "name": table_name,
            "style": table_style,
            "first_column": index,
            "columns": [
                {"header": col_name, "format": format_for_col(df[col], format_mapping)}
                for col, col_name in zip(df.columns, column_names)
            ],
        }
        ws.add_table(0, 0, len(df), len(df.columns) - 1, options)

        if header_orientation == "diagonal":
            ws.set_row(
                0, max(15, 12 + 4 * max(len(c) for c in column_names)), header_format
            )
        elif header_orientation == "vertical":
            ws.set_row(
                0, max(15, 4 + 6 * max(len(c) for c in column_names)), header_format
            )
        elif header_orientation == "horizontal":
            # adjust row widths
            for i, width in enumerate(len(str(x)) for x in column_names):
                ws.set_column(i, i, max(8.43, width))
    wb.close()
    return


def df_to_xlsx_table(
    df: pd.DataFrame,
    table_name: str,
    file: Optional[Union[str, BinaryIO]] = None,
    index: bool = True,
    table_style: Optional[TableStyleInfo] = "Table Style Medium 9",
    nan_inf_to_errors=False,
    header_orientation: HeaderOrientation = "horizontal",
    remove_timezone: bool = False,
) -> None:
    """Convert single dataframe to an excel file.

    Args:
        df (DataFrame): Padas dataframe to convert to excel.
        table_name (str):Name of the table.
        file (Union[str, BinaryIO]): File name or descriptor for the output.
            Defaults to <table_name>.xlsx
        index (bool, optional): Include the datafrme index in the results. Defaults
            to True
        table_style (Optional[NamedTableStyle], optional): Excel table style. Defaults
            to "Table Style Medium 9".
        nan_inf_to_errors (bool, optional): Explicitly write nan/inf values as errors.
            Defaults to False.
        header_orientation (HeaderOrientation, optional): Rotate the table headers, can
            be horizontal, vertical or diagonal. Defaults to "horizontal".
    """
    dfs_to_xlsx_tables(
        [(df, table_name)],
        file=file or table_name + ".xlsx",
        index=index,
        table_style=table_style,
        nan_inf_to_errors=nan_inf_to_errors,
        header_orientation=header_orientation,
        remove_timezone=remove_timezone,
    )

